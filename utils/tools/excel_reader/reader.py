"""
excel_reader/reader.py
======================
Lee cualquier archivo .xlsx y exporta su contenido a un archivo .txt.

Uso:
    python reader.py <ruta_al_excel> [--output <ruta_salida.txt>]

Si no se pasa --output, crea un archivo con el mismo nombre que el Excel
pero con extensión .txt en la misma carpeta.
"""

import argparse
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl no está instalado. Ejecuta: pip install openpyxl")
    sys.exit(1)


def read_excel(file_path: Path, output_path: Path) -> None:
    """Lee todas las hojas del Excel y escribe el contenido en output_path."""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except FileNotFoundError:
        print(f"ERROR: No se encontró el archivo: {file_path}")
        sys.exit(1)
    except Exception as e:
        print(f"ERROR al abrir el archivo: {e}")
        sys.exit(1)

    sheet_names = wb.sheetnames

    with open(output_path, "w", encoding="utf-8") as f:

        # ── Encabezado ──────────────────────────────────────────────────────
        f.write(f"Archivo  : {file_path.name}\n")
        f.write(f"Hojas    : {len(sheet_names)} → {sheet_names}\n")
        f.write("=" * 80 + "\n\n")

        for sheet_name in sheet_names:
            ws = wb[sheet_name]

            # Filas con al menos un valor no nulo
            data_rows = [
                (i + 1, list(row))
                for i, row in enumerate(ws.iter_rows(values_only=True))
                if any(cell is not None for cell in row)
            ]

            f.write(f"╔══ HOJA: '{sheet_name}' " + "═" * max(0, 60 - len(sheet_name)) + "\n")
            f.write(f"║  Dimensiones : {ws.dimensions}\n")
            f.write(f"║  Filas totales: {ws.max_row}  |  Columnas totales: {ws.max_column}\n")
            f.write(f"║  Filas con datos: {len(data_rows)}\n")
            f.write("╚" + "═" * 70 + "\n\n")

            if not data_rows:
                f.write("  (Hoja vacía)\n\n")
                continue

            for row_num, row in data_rows:
                # Limpia valores None para facilitar lectura
                clean = [str(v) if v is not None else "" for v in row]
                f.write(f"  Fila {row_num:>4}: {clean}\n")

            f.write("\n")

    print(f"Listo. Resultado guardado en:\n  {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Lee cualquier archivo .xlsx y exporta su contenido a .txt"
    )
    parser.add_argument(
        "excel",
        nargs="?",
        help="Ruta al archivo .xlsx (obligatorio si no hay valor por defecto)",
    )
    parser.add_argument(
        "--output", "-o",
        help="Ruta del archivo de salida .txt (opcional; por defecto: mismo nombre que el Excel)",
    )

    args = parser.parse_args()

    # ── Resolver ruta del Excel ──────────────────────────────────────────────
    if args.excel:
        excel_path = Path(args.excel)
    else:
        # Fallback: busca el primer .xlsx en la misma carpeta que este script
        script_dir = Path(__file__).parent
        xlsx_files = list(script_dir.glob("*.xlsx"))
        if not xlsx_files:
            print("ERROR: No se indicó un archivo y no hay .xlsx en la carpeta del script.")
            parser.print_help()
            sys.exit(1)
        excel_path = xlsx_files[0]
        print(f"[Auto] Usando: {excel_path.name}")

    # ── Resolver ruta de salida ──────────────────────────────────────────────
    if args.output:
        output_path = Path(args.output)
    else:
        output_path = excel_path.with_suffix(".txt")

    read_excel(excel_path, output_path)


if __name__ == "__main__":
    main()
