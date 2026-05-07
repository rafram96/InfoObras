"""
Generador de Excel con formato Lircay (modelo del cliente).

5 hojas (en este orden):
  1. PROFESIONALES        — listado nominal de profesionales propuestos (Paso 2)
  2. REQUISITOS_TDR       — criterios TDR por cargo (Paso 1)
  3. BD_EXPERIENCIAS      — base de datos de experiencias con alertas (Paso 3)
  4. ANALISIS_RTM         — evaluacion cumple/no cumple por criterio (Paso 4)
  5. RESUMEN              — totales y alertas criticas (al final)

A diferencia de excel_writer.py (formato propio):
- Alertas en lenguaje natural (no codigos ALT01/ALT02)
- Sin hojas separadas de Alertas / Verificacion InfoObras (todo embebido en BD)
- Hojas adicionales para profesionales y requisitos TDR
- Resumen al final, no al inicio

Modulo experimental: NO reemplaza excel_writer.py. Se invoca por separado.
"""
from __future__ import annotations
import logging
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.extraction.models import (
    Experience,
    Professional,
    RequisitoPersonal,
    EvaluacionRTM,
    ResultadoProfesional,
)
from src.validation.rules import Alert, AlertCode, Severity

logger = logging.getLogger(__name__)


# ── Estilos ──────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill(start_color="022448", end_color="022448", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
_CELL_ALIGN_TOP = Alignment(vertical="top", wrap_text=True)

_FILL_VERDE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FILL_AMARILLO = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_FILL_ROJO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_FILL_GRIS = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
_FILL_AZUL_SUAVE = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


# ── Helpers ──────────────────────────────────────────────────────────────────

def _aplicar_header(ws, headers: list[str], row: int = 1) -> None:
    """Estiliza la fila de headers."""
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = _HEADER_ALIGN
        c.border = _BORDER
    ws.row_dimensions[row].height = 38


def _set_column_widths(ws, widths: list[int]) -> None:
    """Setea anchos de columnas."""
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _fmt_date(d) -> str:
    """Formatea date a dd/mm/yyyy."""
    if d is None:
        return ""
    if isinstance(d, str):
        return d
    if isinstance(d, (date, datetime)):
        return d.strftime("%d/%m/%Y")
    return str(d)


def _color_cumple(valor: Optional[str]) -> Optional[PatternFill]:
    """Retorna fill segun el valor (CUMPLE/SI/NO CUMPLE/NO)."""
    if not valor:
        return None
    v = valor.strip().upper()
    if v in ("CUMPLE", "SI", "SÍ"):
        return _FILL_VERDE
    if v.startswith("NO"):
        return _FILL_ROJO
    if "OBSERV" in v or "REVIS" in v:
        return _FILL_AMARILLO
    return None


# ── Clasificacion empresa publica/privada ────────────────────────────────────

_KEYWORDS_PUBLICA = (
    "MINISTERIO", "MINSA", "PRONIS", "ESSALUD", "GOBIERNO REGIONAL",
    "GOBIERNO LOCAL", "MUNICIPALIDAD", "MUNICIPIO", "GOBIERNO PROVINCIAL",
    "MINEDU", "MIDIS", "MINJUS", "MINAGRI", "MINAM", "MTC",
    "PROVIAS", "OINFE", "PNSU", "PNSR", "OFICINA DE INVERSIONES",
    "PROGRAMA NACIONAL", "ENTIDAD PUBLICA", "PROYECTO ESPECIAL",
    "INSTITUTO NACIONAL", "AUTORIDAD NACIONAL", "AUTORIDAD REGIONAL",
    "POLICIA NACIONAL", "ESTADO PERUANO", "EMPRESA PUBLICA", "EPS ",
    "DIRESA", "GERESA", "REDESS", "CENATE", "CENARES", "SISMED",
    "EJERCITO", "MARINA DE GUERRA", "FUERZA AEREA",
    "REGISTRO NACIONAL", "RENIEC", "SUNAT", "SUNAFIL",
    "UNIVERSIDAD NACIONAL",
)


def _clasificar_empresa(company: Optional[str], ruc: Optional[str]) -> str:
    """Heuristica simple: publica vs privada por nombre de la empresa."""
    if not company:
        return ""
    c = company.upper()
    if any(kw in c for kw in _KEYWORDS_PUBLICA):
        return "Pública"
    return "Privada"


# ── Mapeo de alertas a texto natural ─────────────────────────────────────────
# Tomamos el `description` que ya viene de la dataclass Alert (lenguaje natural).
# Estas funciones extraen alertas especificas para columnas dedicadas.

def _alerta_por_codigo(
    alertas: list[Alert],
    codigo: AlertCode,
) -> str:
    """Retorna la description de la primera alerta del codigo dado."""
    for a in alertas:
        if a.code == codigo:
            return a.description or ""
    return ""


def _todas_alertas_naturales(alertas: list[Alert]) -> str:
    """Junta todas las descripciones de alertas en lenguaje natural, separadas por '|'."""
    if not alertas:
        return ""
    descs = [a.description for a in alertas if a.description]
    return " | ".join(descs)


# ── Calculos auxiliares ──────────────────────────────────────────────────────

def _anos_decimal(start: Optional[date], end: Optional[date]) -> Optional[float]:
    """Convierte un periodo a anos decimales (incluyendo dias)."""
    if not start or not end:
        return None
    if not isinstance(start, date) or not isinstance(end, date):
        return None
    delta_dias = (end - start).days
    if delta_dias < 0:
        return None
    return delta_dias / 365.0


def _duracion_human(start: Optional[date], end: Optional[date]) -> str:
    """Retorna texto humano: 'X meses y Y días'."""
    if not start or not end:
        return ""
    if not isinstance(start, date) or not isinstance(end, date):
        return ""
    dias_total = (end - start).days
    if dias_total < 0:
        return ""
    meses = dias_total // 30
    dias = dias_total % 30
    if dias == 0:
        return f"{meses} meses"
    return f"{meses} meses y {dias} días"


def _es_covid_periodo(start: Optional[date], end: Optional[date]) -> bool:
    """True si el periodo intersecta con el COVID (16/03/2020 - 31/12/2021)."""
    if not start or not end:
        return False
    covid_ini = date(2020, 3, 16)
    covid_fin = date(2021, 12, 31)
    return start <= covid_fin and end >= covid_ini


# ============================================================================
# HOJA 1: PROFESIONALES
# ============================================================================

def _write_profesionales(wb, resultados: list[ResultadoProfesional]) -> None:
    """
    Lista nominal de profesionales propuestos.
    Cols: N°, Nombre, Profesión, Fecha colegiación, Especialidad, Folio colegiatura, Folio nombramiento.
    """
    ws = wb.create_sheet("PROFESIONALES")
    headers = [
        "No",
        "Columna 1: Nombre del Profesional",
        "Columna 2: Profesión (según título profesional)",
        "Columna 3: Fecha de colegiación",
        "Columna 4: Especialidad a la que postulan",
        "Columna 5: Folio de la colegiatura",
        "Columna 6: Folio del nombre del profesional",
    ]
    _aplicar_header(ws, headers)
    _set_column_widths(ws, [5, 38, 30, 15, 38, 18, 22])

    for idx, rp in enumerate(resultados, start=1):
        prof = rp.profesional
        ws.cell(row=idx + 1, column=1, value=idx)
        ws.cell(row=idx + 1, column=2, value=(prof.name or "").upper())
        ws.cell(row=idx + 1, column=3, value=(prof.profession or "").upper())
        ws.cell(row=idx + 1, column=4, value=_fmt_date(prof.registration_date))
        ws.cell(row=idx + 1, column=5, value=(prof.role or "").upper())
        # Folio de colegiatura — el campo Professional.folio es general.
        # Por defecto usamos el folio para ambos.
        ws.cell(row=idx + 1, column=6, value=prof.folio or "")
        ws.cell(row=idx + 1, column=7, value=prof.folio or "")

        for col in range(1, 8):
            c = ws.cell(row=idx + 1, column=col)
            c.alignment = _CELL_ALIGN_TOP
            c.border = _BORDER

    ws.freeze_panes = "B2"


# ============================================================================
# HOJA 2: REQUISITOS_TDR
# ============================================================================

def _write_requisitos_tdr(wb, resultados: list[ResultadoProfesional]) -> None:
    """
    Criterios TDR por cargo (Paso 1).
    Cols 1-6: autogeneradas. Cols 7-10: vacias para input manual del evaluador.
    """
    ws = wb.create_sheet("REQUISITOS_TDR")
    headers = [
        "Columna 1: Cargo y Profesión (Titulado Profesional)",
        "Columna 2: Años de Colegiado",
        "Columna 3: Requisito Mínimo (Tiempo y Cargos Similares) y Puntuación",
        "Columna 4: Tipo de experiencia similar (Tipo de obra)",
        "Columna 5: Tiempo adicional en Factores de Evaluación",
        "Columna 6: Capacitación en Factores de Evaluación",
        # Columnas manuales del evaluador (en blanco):
        "Columna 7: Nombre asignado",
        "Columna 8: Trabajará en el cargo",
        "Columna 9: Monto Total",
        "Columna 10: Años Acumulados",
    ]
    _aplicar_header(ws, headers)
    _set_column_widths(ws, [42, 18, 50, 32, 32, 32, 22, 22, 14, 16])

    # Tomamos cada RequisitoPersonal unico (por cargo)
    requisitos_vistos = set()
    fila = 2
    for idx, rp in enumerate(resultados, start=1):
        req = rp.requisito
        if not req or req.cargo in requisitos_vistos:
            continue
        requisitos_vistos.add(req.cargo)

        # Col 1: cargo + profesion
        profs_str = " y/o ".join(req.profesiones_aceptadas or []) or ""
        cargo_profesion = f"{idx}. {req.cargo}"
        if profs_str:
            cargo_profesion += f"\n{profs_str}"
        ws.cell(row=fila, column=1, value=cargo_profesion)

        # Col 2: anos colegiado
        ws.cell(row=fila, column=2, value=req.anos_colegiado or "")

        # Col 3: requisito minimo (tiempo + cargos similares)
        exp_min = req.experiencia_minima
        if exp_min:
            cargos_lista = " y/o ".join(exp_min.cargos_similares_validos or [])
            req_text = f"{exp_min.cantidad or '?'} {exp_min.unidad}."
            if cargos_lista:
                req_text += f"\nCargos: {cargos_lista}"
            if exp_min.puntaje_maximo:
                req_text += f"\nPuntaje: {exp_min.puntaje_maximo}"
            ws.cell(row=fila, column=3, value=req_text)
        else:
            ws.cell(row=fila, column=3, value="")

        # Col 4: tipo de obra
        ws.cell(row=fila, column=4, value=req.tipo_obra_valido or "")

        # Col 5: tiempo adicional
        ws.cell(row=fila, column=5, value=req.tiempo_adicional_factores or "")

        # Col 6: capacitacion
        cap = req.capacitacion or {}
        if isinstance(cap, dict):
            cap_text = cap.get("tema") or ""
            if cap.get("duracion_minima_horas"):
                cap_text += f"\nDuración mínima: {cap['duracion_minima_horas']} h"
        else:
            cap_text = str(cap)
        ws.cell(row=fila, column=6, value=cap_text)

        # Cols 7-10: vacias (input manual del evaluador)

        # Estilos
        for col in range(1, 11):
            c = ws.cell(row=fila, column=col)
            c.alignment = _CELL_ALIGN_TOP
            c.border = _BORDER
            if col >= 7:
                c.fill = _FILL_AZUL_SUAVE  # marca visual de "input manual"

        ws.row_dimensions[fila].height = 60
        fila += 1

    ws.freeze_panes = "B2"


# ============================================================================
# HOJA 3: BD_EXPERIENCIAS
# ============================================================================

def _write_bd_experiencias(
    wb,
    resultados: list[ResultadoProfesional],
    proposal_date: Optional[date] = None,
) -> None:
    """
    Base de datos de todas las experiencias declaradas (Paso 3).
    27 columnas (mismo layout que el Lircay original).
    """
    ws = wb.create_sheet("BD_EXPERIENCIAS")
    headers = [
        "Col 1: Nombre del Profesional",
        "Col 2: DNI / Colegiatura",
        "Col 3: Nombre del proyecto",
        "Col 4: Cargo en el proyecto",
        "Col 5: Consorcio o empresa emisora",
        "Col 6: RUC emisor",
        "Col 7: Pública o Privada",
        "Col 8: Tipo de acreditación",
        "Col 9: Fecha de inicio",
        "Col 10: Fecha de fin",
        "Col 11: Alerta COVID",
        "Col 12: Años decimal",                       # calculo interno
        "Col 13: Años acumulados (profesional)",      # calculo interno
        "Col 14: Duración de la experiencia",
        "Col 15: Fecha de emisión",
        "Col 16: ALERTA emisión",                     # ALT01
        "Col 17: Folio",
        "Col 18: Firma el certificado",
        "Col 19: Cargo del firmante",
        "Col 20: ALERTA no representante",
        "Col 21: Fecha creación (SUNAT)",             # ALT04
        "Col 22: ALERTA creación",                    # ALT04
        "Col 23: ALERTA > 25 años",                   # ALT03 (cliente actualizo de 20 a 25)
        "Col 24: Documento presentado",
        "Col 25: Código CUI / CIU",
        "Col 26: Código InfoObras",
        "Col 27: Fecha creación y ALERTA",
    ]
    _aplicar_header(ws, headers)
    _set_column_widths(
        ws,
        [30, 18, 40, 28, 36, 14, 14, 22, 14, 14, 30, 12, 12, 22, 14, 35,
         12, 28, 22, 35, 14, 35, 35, 22, 16, 16, 22],
    )

    # Calcular acumulados por profesional (suma de anos_decimal)
    acumulados_por_prof: dict[str, float] = {}
    for rp in resultados:
        nombre = rp.profesional.name or ""
        for ev in rp.evaluaciones:
            exp = ev.experiencia_ref
            if not exp:
                continue
            anos = _anos_decimal(exp.start_date, exp.end_date)
            if anos:
                acumulados_por_prof[nombre] = acumulados_por_prof.get(nombre, 0.0) + anos

    fila = 2
    # Para columna 13 (acumulado), solo mostramos en la ULTIMA experiencia del profesional
    contadores_filas: dict[str, int] = {}
    for rp in resultados:
        nombre = rp.profesional.name or ""
        contadores_filas[nombre] = sum(1 for ev in rp.evaluaciones)

    contador_actual: dict[str, int] = {}
    for rp in resultados:
        prof = rp.profesional
        nombre = prof.name or ""
        for ev in rp.evaluaciones:
            exp = ev.experiencia_ref
            if not exp:
                continue
            contador_actual[nombre] = contador_actual.get(nombre, 0) + 1

            alertas = list(ev.alertas)
            anos_dec = _anos_decimal(exp.start_date, exp.end_date)

            # Col 1
            ws.cell(row=fila, column=1, value=(exp.professional_name or nombre).upper())
            # Col 2 — DNI o colegiatura
            dni_col = exp.dni or ""
            if not dni_col and prof.tipo_colegio and prof.registro_colegio:
                dni_col = f"{prof.tipo_colegio} N° {prof.registro_colegio}"
            ws.cell(row=fila, column=2, value=dni_col)
            # Col 3-4
            ws.cell(row=fila, column=3, value=exp.project_name or "")
            ws.cell(row=fila, column=4, value=exp.role or "")
            # Col 5-6
            ws.cell(row=fila, column=5, value=exp.company or "")
            ws.cell(row=fila, column=6, value=exp.ruc or "")
            # Col 7
            ws.cell(row=fila, column=7, value=_clasificar_empresa(exp.company, exp.ruc))
            # Col 8
            ws.cell(row=fila, column=8, value=exp.tipo_acreditacion or "")
            # Col 9-10
            ws.cell(row=fila, column=9, value=_fmt_date(exp.start_date))
            ws.cell(row=fila, column=10, value=_fmt_date(exp.end_date) if exp.end_date else "A LA FECHA")
            # Col 11 — alerta COVID
            covid_alert = _alerta_por_codigo(alertas, AlertCode.PERIODO_COVID)
            ws.cell(row=fila, column=11, value=covid_alert)
            if covid_alert:
                ws.cell(row=fila, column=11).fill = _FILL_AMARILLO
            # Col 12 — años decimal
            if anos_dec is not None:
                ws.cell(row=fila, column=12, value=round(anos_dec, 4))
            # Col 13 — acumulado (solo en la ultima fila de cada profesional)
            if contador_actual.get(nombre) == contadores_filas.get(nombre):
                acum = acumulados_por_prof.get(nombre)
                if acum is not None:
                    ws.cell(row=fila, column=13, value=round(acum, 4))
            # Col 14 — duracion humana
            ws.cell(row=fila, column=14, value=_duracion_human(exp.start_date, exp.end_date))
            # Col 15 — fecha emision
            ws.cell(row=fila, column=15, value=_fmt_date(exp.cert_issue_date))
            # Col 16 — alerta emision (ALT01)
            alt_emision = _alerta_por_codigo(alertas, AlertCode.FIN_DESPUES_EMISION)
            ws.cell(row=fila, column=16, value=alt_emision)
            if alt_emision:
                ws.cell(row=fila, column=16).fill = _FILL_ROJO
            # Col 17 — folio
            ws.cell(row=fila, column=17, value=exp.folio or "")
            # Col 18-19 — firmante
            ws.cell(row=fila, column=18, value=exp.signer or "")
            ws.cell(row=fila, column=19, value=exp.cargo_firmante or "")
            # Col 20 — alerta no representante (heuristica: si cargo_firmante no es Rep Legal/Gerente)
            cargo_f = (exp.cargo_firmante or "").upper()
            es_rep_legal = any(k in cargo_f for k in [
                "REPRESENTANTE LEGAL", "GERENTE GENERAL", "GERENTE", "DIRECTOR",
                "PRESIDENTE", "APODERADO",
            ])
            if cargo_f and not es_rep_legal:
                ws.cell(
                    row=fila, column=20,
                    value="El firmante no aparece como representante legal de la empresa emisora",
                )
                ws.cell(row=fila, column=20).fill = _FILL_AMARILLO
            # Col 21 — fecha SUNAT (no integrada todavia)
            ws.cell(row=fila, column=21, value="")
            # Col 22 — alerta creacion (ALT04)
            alt_creacion = _alerta_por_codigo(alertas, AlertCode.EMPRESA_POST_EXPERIENCIA)
            ws.cell(row=fila, column=22, value=alt_creacion)
            if alt_creacion:
                ws.cell(row=fila, column=22).fill = _FILL_AMARILLO
            # Col 23 — alerta > 25 anos (ALT03)
            alt_25 = _alerta_por_codigo(alertas, AlertCode.MAS_25_ANOS)
            ws.cell(row=fila, column=23, value=alt_25)
            if alt_25:
                ws.cell(row=fila, column=23).fill = _FILL_ROJO
            # Col 24 — documento presentado
            ws.cell(row=fila, column=24, value=exp.tipo_acreditacion or "")
            # Col 25-26 — CUI e InfoObras
            ws.cell(row=fila, column=25, value=exp.cui or "")
            ws.cell(row=fila, column=26, value=exp.infoobras_code or "")
            # Col 27 — fecha creacion + alerta combinada (vacio por ahora)
            ws.cell(row=fila, column=27, value="")

            # Estilos
            for col in range(1, 28):
                c = ws.cell(row=fila, column=col)
                c.alignment = _CELL_ALIGN_TOP
                c.border = _BORDER
            ws.row_dimensions[fila].height = 50
            fila += 1

    ws.freeze_panes = "B2"


# ============================================================================
# HOJA 4: ANALISIS_RTM
# ============================================================================

def _write_analisis_rtm(wb, resultados: list[ResultadoProfesional]) -> None:
    """
    Evaluacion cumple/no cumple por criterio (Paso 4). 25 cols.
    """
    ws = wb.create_sheet("ANALISIS_RTM")
    headers = [
        "Col 1: Cargo en el proyecto o OBRA",
        "Col 2: Nombre del Profesional",
        "Col 3: Profesión propuesta",
        "Col 4: Profesión indicada en las bases",
        "Col 5: ¿Cumple la profesión?",
        "Col 6: Folio del certificado",
        "Col 7: Cargo de las experiencias propuestas",
        "Col 8: Cargos válidos según bases",
        "Col 9: ¿Cumple el cargo?",
        "Col 10: Actividades señaladas en el certificado",
        "Col 11: Cargos válidos según bases",          # repetido intencional
        "Col 12: ¿Actividades coinciden con el cargo?",
        "Col 13: Proyecto o experiencia propuesta",
        "Col 14: Obra o proyecto válido según bases",
        "Col 15: ¿Cumple el proyecto/obra?",
        "Col 16: Fecha de término",
        "Col 17: ¿No indica fecha o dice ACTUALIDAD?",
        "Col 18: Tipo de obra (certificado)",
        "Col 19: Tipos de obra solicitados (bases)",
        "Col 20: ¿Cumple tipo de obra?",
        "Col 21: Tipo de intervención (certificado)",
        "Col 22: Tipo de intervención (bases)",
        "Col 23: ¿Cumple tipo de intervención?",
        "Col 24: Acredita complejidad",
        "Col 25: ¿Término dentro de los últimos 25 años?",
    ]
    _aplicar_header(ws, headers)
    _set_column_widths(
        ws,
        [30, 30, 24, 40, 14, 12, 28, 40, 14, 35, 40, 18,
         40, 35, 14, 14, 14, 30, 35, 14, 30, 30, 14, 14, 14],
    )

    fila = 2
    for rp in resultados:
        for ev in rp.evaluaciones:
            exp = ev.experiencia_ref

            # Col 1 — cargo postulado
            ws.cell(row=fila, column=1, value=(ev.cargo_postulado or rp.profesional.role or "").upper())
            # Col 2 — nombre
            ws.cell(row=fila, column=2, value=(ev.nombre or rp.profesional.name or "").upper())
            # Col 3 — profesion propuesta
            ws.cell(row=fila, column=3, value=(ev.profesion_propuesta or "").upper())
            # Col 4 — profesion en bases
            ws.cell(row=fila, column=4, value=ev.profesion_requerida or "")
            # Col 5 — cumple profesion
            ws.cell(row=fila, column=5, value=ev.cumple_profesion or "")
            f = _color_cumple(ev.cumple_profesion)
            if f:
                ws.cell(row=fila, column=5).fill = f
            # Col 6 — folio
            ws.cell(row=fila, column=6, value=ev.folio_certificado or "")
            # Col 7 — cargo experiencia
            ws.cell(row=fila, column=7, value=ev.cargo_experiencia or "")
            # Col 8 — cargos validos bases
            ws.cell(row=fila, column=8, value=ev.cargos_validos_bases or "")
            # Col 9 — cumple cargo
            ws.cell(row=fila, column=9, value=ev.cumple_cargo or "")
            f = _color_cumple(ev.cumple_cargo)
            if f:
                ws.cell(row=fila, column=9).fill = f

            # Col 10 — actividades del certificado (heuristica)
            actividades = "NO PRECISA ACTIVIDADES O FUNCIONES"
            if exp and exp.raw_text:
                # Si el raw_text del certificado tiene "actividades"/"funciones" + parrafo
                rt = exp.raw_text.lower()
                if any(kw in rt for kw in ["actividades:", "funciones:", "tareas:",
                                              "realizo", "realizó", "encargado de",
                                              "responsable de"]):
                    actividades = exp.raw_text[:300]
            ws.cell(row=fila, column=10, value=actividades)
            # Col 11 — cargos validos bases (REPETIDO para comparacion)
            ws.cell(row=fila, column=11, value=ev.cargos_validos_bases or "")
            # Col 12 — actividades coinciden
            actividades_coinciden = "NO HAY ACTIVIDADES O FUNCIONES"
            if actividades != "NO PRECISA ACTIVIDADES O FUNCIONES":
                # Si tiene actividades, usar el resultado del cumple_cargo como proxy
                actividades_coinciden = ev.cumple_cargo or "NO EVALUABLE"
            ws.cell(row=fila, column=12, value=actividades_coinciden)
            f = _color_cumple(actividades_coinciden)
            if f:
                ws.cell(row=fila, column=12).fill = f

            # Col 13 — proyecto propuesto
            ws.cell(row=fila, column=13, value=ev.proyecto_propuesto or "")
            # Col 14 — proyecto valido bases
            ws.cell(row=fila, column=14, value=ev.proyecto_valido_bases or "")
            # Col 15 — cumple proyecto
            ws.cell(row=fila, column=15, value=ev.cumple_proyecto or "")
            f = _color_cumple(ev.cumple_proyecto)
            if f:
                ws.cell(row=fila, column=15).fill = f

            # Col 16 — fecha termino
            ws.cell(row=fila, column=16, value=_fmt_date(ev.fecha_termino))
            # Col 17 — sin fecha o ACTUALIDAD
            sin_fecha = ev.alerta_fecha_termino or ("SI" if not ev.fecha_termino else "NO")
            ws.cell(row=fila, column=17, value=sin_fecha)
            if "SI" in sin_fecha or "NO VALE" in sin_fecha:
                ws.cell(row=fila, column=17).fill = _FILL_AMARILLO

            # Col 18-20 — tipo obra
            ws.cell(row=fila, column=18, value=ev.tipo_obra_certificado or "")
            ws.cell(row=fila, column=19, value=ev.tipo_obra_requerido or "")
            ws.cell(row=fila, column=20, value=ev.cumple_tipo_obra or "")
            f = _color_cumple(ev.cumple_tipo_obra)
            if f:
                ws.cell(row=fila, column=20).fill = f

            # Col 21-23 — tipo intervencion
            ws.cell(row=fila, column=21, value=ev.intervencion_certificado or "")
            ws.cell(row=fila, column=22, value=ev.intervencion_requerida or "")
            ws.cell(row=fila, column=23, value=ev.cumple_intervencion or "")
            f = _color_cumple(ev.cumple_intervencion)
            if f:
                ws.cell(row=fila, column=23).fill = f

            # Col 24 — acredita complejidad
            ws.cell(row=fila, column=24, value=ev.acredita_complejidad or "")
            f = _color_cumple(ev.acredita_complejidad)
            if f:
                ws.cell(row=fila, column=24).fill = f

            # Col 25 — dentro de 25 anos
            ws.cell(row=fila, column=25, value=ev.dentro_20_anos or "")
            f = _color_cumple(ev.dentro_20_anos)
            if f:
                ws.cell(row=fila, column=25).fill = f

            # Estilos
            for col in range(1, 26):
                c = ws.cell(row=fila, column=col)
                c.alignment = _CELL_ALIGN_TOP
                c.border = _BORDER
            ws.row_dimensions[fila].height = 60
            fila += 1

    ws.freeze_panes = "B2"


# ============================================================================
# HOJA 5: RESUMEN (al final)
# ============================================================================

def _write_resumen(
    wb,
    resultados: list[ResultadoProfesional],
    proposal_date: Optional[date] = None,
    filename: str = "",
) -> None:
    """Resumen ejecutivo al final."""
    ws = wb.create_sheet("RESUMEN")
    _set_column_widths(ws, [4, 35, 70])

    # Titulo
    c = ws.cell(row=1, column=1, value="RESUMEN DE EVALUACIÓN")
    c.font = Font(name="Calibri", size=16, bold=True, color="022448")
    ws.merge_cells("A1:C1")

    # Subtitulo
    c = ws.cell(row=2, column=1, value=filename or "Propuesta técnica")
    c.font = Font(name="Calibri", size=11, italic=True)
    ws.merge_cells("A2:C2")

    fila = 4

    def _info(label: str, valor):
        nonlocal fila
        ws.cell(row=fila, column=2, value=label).font = Font(bold=True)
        ws.cell(row=fila, column=3, value=str(valor))
        fila += 1

    _info("Fecha de propuesta", _fmt_date(proposal_date))
    _info("Total de profesionales", len(resultados))

    total_experiencias = sum(len(rp.evaluaciones) for rp in resultados)
    _info("Total de experiencias evaluadas", total_experiencias)

    # Conteos por severidad
    severidades = {Severity.CRITICA: 0, Severity.OBSERVACION: 0, Severity.INFORMATIVA: 0}
    for rp in resultados:
        for ev in rp.evaluaciones:
            for a in ev.alertas:
                severidades[a.severity] = severidades.get(a.severity, 0) + 1
        for a in rp.alertas_globales:
            severidades[a.severity] = severidades.get(a.severity, 0) + 1
    _info("Alertas críticas", severidades.get(Severity.CRITICA, 0))
    _info("Alertas de observación", severidades.get(Severity.OBSERVACION, 0))
    _info("Alertas informativas", severidades.get(Severity.INFORMATIVA, 0))

    # Cumplimiento global
    total_eval = total_experiencias
    cumplen_profesion = sum(
        1 for rp in resultados for ev in rp.evaluaciones
        if (ev.cumple_profesion or "").upper().startswith("SI")
    )
    cumplen_cargo = sum(
        1 for rp in resultados for ev in rp.evaluaciones
        if (ev.cumple_cargo or "").upper().startswith("CUMPLE")
    )
    if total_eval > 0:
        _info("% Cumple profesión", f"{cumplen_profesion / total_eval:.0%}")
        _info("% Cumple cargo", f"{cumplen_cargo / total_eval:.0%}")

    fila += 1
    # Lista de profesionales con resumen 1-linea
    c = ws.cell(row=fila, column=2, value="Profesionales evaluados")
    c.font = Font(bold=True, size=12)
    fila += 1
    for idx, rp in enumerate(resultados, start=1):
        nombre = rp.profesional.name or "(sin nombre)"
        cargo = rp.profesional.role or "(sin cargo)"
        n_alertas_criticas = sum(
            1 for ev in rp.evaluaciones for a in ev.alertas
            if a.severity == Severity.CRITICA
        )
        marca = "🔴" if n_alertas_criticas > 0 else "✅"
        ws.cell(row=fila, column=2, value=f"{idx}. {nombre}")
        ws.cell(row=fila, column=3, value=f"{cargo} — {marca} {n_alertas_criticas} alertas críticas")
        fila += 1


# ============================================================================
# Funcion principal
# ============================================================================

def write_report_lircay(
    resultados: list[ResultadoProfesional],
    output_path: Path,
    proposal_date: Optional[date] = None,
    filename: str = "",
) -> Path:
    """
    Genera el Excel con formato Lircay (5 hojas).

    Args:
        resultados: lista de ResultadoProfesional del Paso 4 (motor de reglas).
        output_path: ruta donde guardar el .xlsx.
        proposal_date: fecha de presentacion de la propuesta.
        filename: nombre del archivo PDF original (para mostrar en RESUMEN).

    Returns:
        Path al archivo Excel generado.

    NOTA: este modulo es experimental, paralelo a excel_writer.py. No reemplaza
    el writer original. Para cambiar el formato de salida del API hay que
    enchufarlo en src/api/main.py donde se invoca write_report().
    """
    wb = openpyxl.Workbook()
    # Borrar la hoja por defecto que crea openpyxl
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _write_profesionales(wb, resultados)
    _write_requisitos_tdr(wb, resultados)
    _write_bd_experiencias(wb, resultados, proposal_date)
    _write_analisis_rtm(wb, resultados)
    _write_resumen(wb, resultados, proposal_date, filename)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    logger.info(
        "[excel-lircay] Reporte generado: %s (%d profesionales, %d experiencias)",
        output_path, len(resultados),
        sum(len(rp.evaluaciones) for rp in resultados),
    )
    return output_path
