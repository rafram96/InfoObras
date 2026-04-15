"""
Genera el Excel final con 5 hojas usando openpyxl.
Colores: Verde = Cumple · Amarillo = Observación · Rojo = No cumple/Alerta crítica
"""
from datetime import date
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.extraction.models import (
    Professional, Experience, EvaluacionRTM, ResultadoProfesional,
)
from src.validation.rules import Alert, Severity, calculate_effective_days, calculate_effective_years

# ── Estilos ──────────────────────────────────────────────────────────────────
GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")
HEADER_FILL = PatternFill("solid", fgColor="022448")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(size=9)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def _apply_header(ws, headers: list[str], row: int = 1) -> None:
    """Aplica estilos de cabecera a una fila."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN_BORDER


def _auto_width(ws, min_width: int = 8, max_width: int = 40) -> None:
    """Ajusta el ancho de columnas basado en contenido."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = min_width
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[col_letter].width = max_len + 2


def _fmt_date(d: Optional[date]) -> str:
    """Formatea date a DD/MM/YYYY o cadena vacía."""
    return d.strftime("%d/%m/%Y") if d else ""


def _cumple_fill(valor: str) -> Optional[PatternFill]:
    """Retorna el fill según el valor de cumplimiento."""
    v = valor.upper() if valor else ""
    if v in ("SI", "CUMPLE"):
        return GREEN
    elif v in ("NO", "NO CUMPLE"):
        return RED
    elif v == "NO EVALUABLE":
        return YELLOW
    return None


# ── Hoja 1: Resumen ─────────────────────────────────────────────────────────

def _write_resumen(
    wb: openpyxl.Workbook,
    resultados: list[ResultadoProfesional],
    proposal_date: Optional[date] = None,
    filename: str = "",
) -> None:
    """Hoja Resumen con totales y alertas críticas."""
    ws = wb.active
    ws.title = "Resumen"

    # Título
    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = "RESUMEN DE EVALUACIÓN"
    cell.font = Font(bold=True, size=14, color="022448")

    # Metadata
    ws["A3"] = "Archivo:"
    ws["B3"] = filename
    ws["A4"] = "Fecha propuesta:"
    ws["B4"] = _fmt_date(proposal_date)
    ws["A5"] = "Fecha análisis:"
    ws["B5"] = _fmt_date(date.today())

    # Totales
    total_prof = len(resultados)
    con_rtm = sum(1 for r in resultados if r.requisito_encontrado)
    total_exp = sum(len(r.evaluaciones) for r in resultados)
    total_alertas = sum(len(ev.alertas) for r in resultados for ev in r.evaluaciones)
    alertas_criticas = sum(
        1 for r in resultados for ev in r.evaluaciones
        for a in ev.alertas if a.severity == Severity.CRITICAL
    )

    ws["A7"] = "Profesionales evaluados:"
    ws["B7"] = total_prof
    ws["A8"] = "Con match RTM:"
    ws["B8"] = f"{con_rtm}/{total_prof}"
    ws["A9"] = "Total experiencias:"
    ws["B9"] = total_exp
    ws["A10"] = "Total alertas:"
    ws["B10"] = total_alertas
    ws["A11"] = "Alertas críticas:"
    ws["B11"] = alertas_criticas
    ws["B11"].fill = RED if alertas_criticas > 0 else GREEN

    # Tabla resumen por profesional
    headers = ["#", "Cargo", "Nombre", "RTM Match", "Experiencias", "Años Efectivos", "Alertas", "Críticas"]
    _apply_header(ws, headers, row=13)

    for i, res in enumerate(resultados, 1):
        row = 13 + i
        n_alertas = sum(len(ev.alertas) for ev in res.evaluaciones)
        n_criticas = sum(
            1 for ev in res.evaluaciones for a in ev.alertas
            if a.severity == Severity.CRITICAL
        )

        # Calcular años efectivos del profesional
        exps_con_fecha = [
            ev.experiencia_ref for ev in res.evaluaciones
            if ev.experiencia_ref and ev.experiencia_ref.start_date and ev.experiencia_ref.end_date
        ]
        anos = calculate_effective_years(exps_con_fecha, proposal_date) if exps_con_fecha and proposal_date else 0.0

        ws.cell(row=row, column=1, value=i).font = BODY_FONT
        ws.cell(row=row, column=2, value=res.profesional.role).font = BODY_FONT
        ws.cell(row=row, column=3, value=res.profesional.name).font = BODY_FONT

        rtm_cell = ws.cell(row=row, column=4, value="SI" if res.requisito_encontrado else "NO")
        rtm_cell.fill = GREEN if res.requisito_encontrado else RED
        rtm_cell.font = BODY_FONT

        ws.cell(row=row, column=5, value=len(res.evaluaciones)).font = BODY_FONT
        ws.cell(row=row, column=6, value=anos).font = BODY_FONT

        ws.cell(row=row, column=7, value=n_alertas).font = BODY_FONT

        crit_cell = ws.cell(row=row, column=8, value=n_criticas)
        crit_cell.font = BODY_FONT
        if n_criticas > 0:
            crit_cell.fill = RED

    _auto_width(ws)


# ── Hoja 2: Base de Datos (27 columnas del Paso 3) ──────────────────────────

def _covid_check(start: Optional[date], end: Optional[date]) -> str:
    """Retorna 'INCLUYE PERIODO COVID' si el periodo solapa con COVID."""
    from src.validation.rules import COVID_START, COVID_END
    if start and end and start <= COVID_END and end >= COVID_START:
        return "INCLUYE PERIODO COVID"
    return ""


def _calc_duracion(start: Optional[date], end: Optional[date]) -> str:
    """Calcula duración en meses entre dos fechas."""
    if not start or not end:
        return ""
    dias = (end - start).days
    if dias <= 0:
        return ""
    meses = dias / 30.44  # promedio de días por mes
    if meses < 1:
        return f"{dias} días"
    return f"{meses:.0f} meses"


def _write_base_datos(
    wb: openpyxl.Workbook,
    resultados: list[ResultadoProfesional],
    proposal_date: Optional[date] = None,
) -> None:
    """Hoja Base de Datos — 27 columnas por experiencia con datos completos."""
    ws = wb.create_sheet("Base de Datos")

    headers = [
        "Nombre Profesional",        # 1
        "DNI / Colegiatura",         # 2
        "Nombre del Proyecto",       # 3
        "Cargo en el Proyecto",      # 4
        "Empresa/Consorcio Emisor",  # 5
        "RUC del Emisor",            # 6
        "Tipo de Obra",              # 7
        "Tipo de Acreditación",      # 8
        "Fecha de Inicio",           # 9
        "Fecha de Fin",              # 10
        "Periodo COVID",             # 11
        "(Reservada)",               # 12
        "(Reservada)",               # 13
        "Duración",                  # 14
        "Fecha de Emisión",          # 15
        "Alerta Emisión",            # 16
        "Folio",                     # 17
        "Nombre del Firmante",       # 18
        "Cargo del Firmante",        # 19
        "Alerta Firmante",           # 20
        "Fecha Creación Emisor",     # 21
        "Alerta Antigüedad Emisor",  # 22
        "Alerta Exp. Antigua",       # 23
        "Tipo de Documento",         # 24
        "Código CUI",               # 25
        "Código InfoObras",          # 26
        "Validación Cruzada Emisor", # 27
    ]
    _apply_header(ws, headers)

    row = 2
    for res in resultados:
        prof = res.profesional
        for ev in res.evaluaciones:
            exp = ev.experiencia_ref  # Experience original

            # Col 1-2: Profesional
            ws.cell(row=row, column=1, value=prof.name).font = BODY_FONT
            ws.cell(row=row, column=2, value=prof.registro_colegio or (exp.dni if exp else "")).font = BODY_FONT

            # Col 3-4: Proyecto y cargo
            ws.cell(row=row, column=3, value=ev.proyecto_propuesto).font = BODY_FONT
            ws.cell(row=row, column=4, value=ev.cargo_experiencia).font = BODY_FONT

            # Col 5-6: Empresa y RUC (de Experience)
            ws.cell(row=row, column=5, value=exp.company if exp else "").font = BODY_FONT
            ws.cell(row=row, column=6, value=exp.ruc if exp else "").font = BODY_FONT

            # Col 7-8: Tipo obra y acreditación
            ws.cell(row=row, column=7, value=ev.tipo_obra_certificado or (exp.tipo_obra if exp else "")).font = BODY_FONT
            ws.cell(row=row, column=8, value=exp.tipo_acreditacion if exp else "").font = BODY_FONT

            # Col 9-10: Fechas (de Experience)
            start = exp.start_date if exp else None
            end = exp.end_date if exp else ev.fecha_termino
            ws.cell(row=row, column=9, value=_fmt_date(start)).font = BODY_FONT
            ws.cell(row=row, column=10, value=_fmt_date(end)).font = BODY_FONT

            # Col 11: COVID check
            covid_txt = _covid_check(start, end)
            covid_cell = ws.cell(row=row, column=11, value=covid_txt)
            covid_cell.font = BODY_FONT
            if covid_txt:
                covid_cell.fill = YELLOW

            # Col 12-13: Reservadas
            ws.cell(row=row, column=12, value="").font = BODY_FONT
            ws.cell(row=row, column=13, value="").font = BODY_FONT

            # Col 14: Duración
            ws.cell(row=row, column=14, value=_calc_duracion(start, end)).font = BODY_FONT

            # Col 15: Fecha emisión
            cert_date = exp.cert_issue_date if exp else None
            ws.cell(row=row, column=15, value=_fmt_date(cert_date)).font = BODY_FONT

            # Col 16: Alerta emisión (fecha emisión < fecha fin → ALERTA)
            alerta_emision = ""
            if cert_date and end and end > cert_date:
                alerta_emision = "ALERTA"
            alerta_em_cell = ws.cell(row=row, column=16, value=alerta_emision)
            alerta_em_cell.font = BODY_FONT
            if alerta_emision:
                alerta_em_cell.fill = YELLOW

            # Col 17: Folio
            ws.cell(row=row, column=17, value=ev.folio_certificado).font = BODY_FONT

            # Col 18-19: Firmante (de Experience)
            ws.cell(row=row, column=18, value=exp.signer if exp else "").font = BODY_FONT
            ws.cell(row=row, column=19, value="").font = BODY_FONT  # cargo_firmante no está en Experience

            # Col 20: Alerta firmante (requiere validación manual)
            ws.cell(row=row, column=20, value="").font = BODY_FONT

            # Col 21: Fecha creación emisor (requiere SUNAT)
            ws.cell(row=row, column=21, value="").font = BODY_FONT

            # Col 22: Alerta antigüedad emisor (requiere SUNAT)
            ws.cell(row=row, column=22, value="").font = BODY_FONT

            # Col 23: Alerta experiencia antigua (>20 años)
            alerta_antigua = ""
            if end and proposal_date:
                try:
                    limite = proposal_date.replace(year=proposal_date.year - 20)
                except ValueError:
                    limite = proposal_date.replace(year=proposal_date.year - 20, day=28)
                if end < limite:
                    alerta_antigua = "ALERTA"
            alerta_ant_cell = ws.cell(row=row, column=23, value=alerta_antigua)
            alerta_ant_cell.font = BODY_FONT
            if alerta_antigua:
                alerta_ant_cell.fill = YELLOW

            # Col 24: Tipo de documento
            ws.cell(row=row, column=24, value=exp.tipo_acreditacion if exp else "").font = BODY_FONT

            # Col 25-26: CUI e InfoObras
            ws.cell(row=row, column=25, value=exp.cui if exp else "").font = BODY_FONT
            ws.cell(row=row, column=26, value=exp.infoobras_code if exp else "").font = BODY_FONT

            # Col 27: Validación cruzada emisor (= col 21, requiere SUNAT)
            ws.cell(row=row, column=27, value="").font = BODY_FONT

            # Borders
            for c in range(1, 28):
                ws.cell(row=row, column=c).border = THIN_BORDER

            row += 1

    _auto_width(ws)


# ── Hoja 3: Evaluación RTM (22 columnas del Paso 4) ─────────────────────────

def _write_evaluacion_rtm(
    wb: openpyxl.Workbook,
    resultados: list[ResultadoProfesional],
) -> None:
    """Hoja Evaluación RTM — 22 columnas por evaluación."""
    ws = wb.create_sheet("Evaluación RTM")

    headers = [
        "Cargo Postulado",              # 1
        "Nombre Profesional",           # 2
        "Profesión Propuesta",          # 3
        "Profesión Requerida",          # 4
        "¿Cumple Profesión?",           # 5
        "Folio Certificado",            # 6
        "Cargo en la Experiencia",      # 7
        "Cargos Válidos (Bases)",       # 8
        "¿Cumple Cargo?",              # 9
        "Proyecto Propuesto",           # 10
        "Proyecto Válido (Bases)",      # 11
        "¿Cumple Proyecto?",           # 12
        "Fecha de Término",             # 13
        "Alerta Fecha Término",         # 14
        "Tipo Obra (Certificado)",      # 15
        "Tipo Obra (Requerido)",        # 16
        "¿Cumple Tipo Obra?",          # 17
        "Intervención (Certificado)",   # 18
        "Intervención (Requerida)",     # 19
        "¿Cumple Intervención?",       # 20
        "¿Acredita Complejidad?",      # 21
        "¿Dentro de 20 Años?",        # 22
    ]
    _apply_header(ws, headers)

    row = 2
    for res in resultados:
        for ev in res.evaluaciones:
            ws.cell(row=row, column=1, value=ev.cargo_postulado).font = BODY_FONT
            ws.cell(row=row, column=2, value=ev.nombre).font = BODY_FONT
            ws.cell(row=row, column=3, value=ev.profesion_propuesta).font = BODY_FONT
            ws.cell(row=row, column=4, value=ev.profesion_requerida).font = BODY_FONT

            c5 = ws.cell(row=row, column=5, value=ev.cumple_profesion)
            c5.font = BODY_FONT
            c5.fill = _cumple_fill(ev.cumple_profesion) or PatternFill()

            ws.cell(row=row, column=6, value=ev.folio_certificado).font = BODY_FONT
            ws.cell(row=row, column=7, value=ev.cargo_experiencia).font = BODY_FONT
            ws.cell(row=row, column=8, value=ev.cargos_validos_bases).font = BODY_FONT

            c9 = ws.cell(row=row, column=9, value=ev.cumple_cargo)
            c9.font = BODY_FONT
            c9.fill = _cumple_fill(ev.cumple_cargo) or PatternFill()

            ws.cell(row=row, column=10, value=ev.proyecto_propuesto).font = BODY_FONT
            ws.cell(row=row, column=11, value=ev.proyecto_valido_bases).font = BODY_FONT

            c12 = ws.cell(row=row, column=12, value=ev.cumple_proyecto)
            c12.font = BODY_FONT
            c12.fill = _cumple_fill(ev.cumple_proyecto) or PatternFill()

            ws.cell(row=row, column=13, value=_fmt_date(ev.fecha_termino)).font = BODY_FONT

            c14 = ws.cell(row=row, column=14, value=ev.alerta_fecha_termino)
            c14.font = BODY_FONT
            if ev.alerta_fecha_termino == "NO VALE":
                c14.fill = RED

            ws.cell(row=row, column=15, value=ev.tipo_obra_certificado).font = BODY_FONT
            ws.cell(row=row, column=16, value=ev.tipo_obra_requerido).font = BODY_FONT

            c17 = ws.cell(row=row, column=17, value=ev.cumple_tipo_obra)
            c17.font = BODY_FONT
            c17.fill = _cumple_fill(ev.cumple_tipo_obra) or PatternFill()

            ws.cell(row=row, column=18, value=ev.intervencion_certificado).font = BODY_FONT
            ws.cell(row=row, column=19, value=ev.intervencion_requerida).font = BODY_FONT

            c20 = ws.cell(row=row, column=20, value=ev.cumple_intervencion)
            c20.font = BODY_FONT
            c20.fill = _cumple_fill(ev.cumple_intervencion) or PatternFill()

            c21 = ws.cell(row=row, column=21, value=ev.acredita_complejidad)
            c21.font = BODY_FONT
            c21.fill = _cumple_fill(ev.acredita_complejidad) or PatternFill()

            c22 = ws.cell(row=row, column=22, value=ev.dentro_20_anos)
            c22.font = BODY_FONT
            c22.fill = _cumple_fill(ev.dentro_20_anos) or PatternFill()

            # Borders
            for c in range(1, 23):
                ws.cell(row=row, column=c).border = THIN_BORDER

            row += 1

    _auto_width(ws)


# ── Hoja 4: Alertas ─────────────────────────────────────────────────────────

def _write_alertas(
    wb: openpyxl.Workbook,
    resultados: list[ResultadoProfesional],
) -> None:
    """Hoja Alertas — todas las alertas generadas."""
    ws = wb.create_sheet("Alertas")

    headers = ["Profesional", "Cargo", "Proyecto", "Código", "Severidad", "Descripción"]
    _apply_header(ws, headers)

    row = 2
    for res in resultados:
        for ev in res.evaluaciones:
            for alerta in ev.alertas:
                ws.cell(row=row, column=1, value=res.profesional.name).font = BODY_FONT
                ws.cell(row=row, column=2, value=res.profesional.role).font = BODY_FONT
                ws.cell(row=row, column=3, value=ev.proyecto_propuesto).font = BODY_FONT
                ws.cell(row=row, column=4, value=alerta.code.value).font = BODY_FONT

                sev_cell = ws.cell(row=row, column=5, value=alerta.severity.value)
                sev_cell.font = BODY_FONT
                sev_cell.fill = RED if alerta.severity == Severity.CRITICAL else YELLOW

                ws.cell(row=row, column=6, value=alerta.description).font = BODY_FONT

                for c in range(1, 7):
                    ws.cell(row=row, column=c).border = THIN_BORDER
                    ws.cell(row=row, column=c).alignment = WRAP

                row += 1

    if row == 2:
        ws.cell(row=2, column=1, value="Sin alertas").font = BODY_FONT

    _auto_width(ws)


# ── Hoja 5: Verificación InfoObras ───────────────────────────────────────────

def _write_infoobras(
    wb: openpyxl.Workbook,
    infoobras_data: Optional[list[dict]] = None,
) -> None:
    """Hoja Verificación InfoObras — datos de obras consultadas."""
    ws = wb.create_sheet("Verificación InfoObras")

    headers = [
        "Profesional", "Proyecto (Certificado)", "CUI",
        "Obra InfoObras", "Estado", "Fecha Inicio Obra",
        "Supervisores", "Residentes", "Paralizaciones",
        "Días Suspensión",
    ]
    _apply_header(ws, headers)

    if not infoobras_data:
        ws.cell(row=2, column=1, value="Sin datos de InfoObras (scrapers no ejecutados)").font = BODY_FONT
        _auto_width(ws)
        return

    row = 2
    for item in infoobras_data:
        ws.cell(row=row, column=1, value=item.get("profesional", "")).font = BODY_FONT
        ws.cell(row=row, column=2, value=item.get("proyecto_cert", "")).font = BODY_FONT
        ws.cell(row=row, column=3, value=item.get("cui", "")).font = BODY_FONT
        ws.cell(row=row, column=4, value=item.get("obra_nombre", "")).font = BODY_FONT
        ws.cell(row=row, column=5, value=item.get("estado", "")).font = BODY_FONT
        ws.cell(row=row, column=6, value=item.get("fecha_inicio", "")).font = BODY_FONT
        ws.cell(row=row, column=7, value=item.get("supervisores", "")).font = BODY_FONT
        ws.cell(row=row, column=8, value=item.get("residentes", "")).font = BODY_FONT
        ws.cell(row=row, column=9, value=item.get("paralizaciones", 0)).font = BODY_FONT

        dias_cell = ws.cell(row=row, column=10, value=item.get("dias_suspension", 0))
        dias_cell.font = BODY_FONT
        if item.get("dias_suspension", 0) > 0:
            dias_cell.fill = YELLOW

        for c in range(1, 11):
            ws.cell(row=row, column=c).border = THIN_BORDER

        row += 1

    _auto_width(ws)


# ── Función principal ────────────────────────────────────────────────────────

def write_report(
    resultados: list[ResultadoProfesional],
    output_path: Path,
    proposal_date: Optional[date] = None,
    filename: str = "",
    infoobras_data: Optional[list[dict]] = None,
) -> Path:
    """
    Genera el Excel de salida con 5 hojas.

    Args:
        resultados: lista de ResultadoProfesional del Paso 4
        output_path: ruta donde guardar el Excel
        proposal_date: fecha de presentación de la propuesta
        filename: nombre del archivo PDF procesado
        infoobras_data: datos de verificación InfoObras (opcional)

    Returns:
        Path al archivo Excel generado
    """
    wb = openpyxl.Workbook()

    _write_resumen(wb, resultados, proposal_date, filename)
    _write_base_datos(wb, resultados, proposal_date)
    _write_evaluacion_rtm(wb, resultados)
    _write_alertas(wb, resultados)
    _write_infoobras(wb, infoobras_data)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return output_path
